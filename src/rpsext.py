from Tkinter import *
from PIL import Image, ImageTk
from openpyxl import load_workbook
import tkFont
import random

width = 800
height = 600

### Openpyxl ###

def getWorksheet(path):
    wb = load_workbook(path)
    ws = wb.active
    return ws

### windows ###

class Application(Frame):

    def __init__(self, parent):
        Frame.__init__(self, parent)

        self.parent = parent
        self.initUI()

    def initUI(self):
        # create font
        customFont = tkFont.Font(window, family="Helvetica", size=18)

        # widgets
        label = Label(window, text="Welcome to RPS: extended!", font=customFont)
        label.pack(fill=X, pady=10)

        button1 = Button(window, text="classic RPS", font=customFont, bg="lightgray", command = lambda: self.rpsWindow())
        button1.pack(fill=X, padx=180, pady=10)

        button2 = Button(window, text="Rock, Paper, Scissors, Spock, Lizard", font=customFont, bg="lightgray", command = lambda: self.rpsslWindow())
        button2.pack(fill=X, padx=180, pady=10)

        button3 = Button(window, text="RPS 7", font=customFont, bg="lightgray")
        button3.pack(fill=X, padx=180, pady=10)

        button4 = Button(window, text="RPS 9", font=customFont, bg="lightgray")
        button4.pack(fill=X, padx=180, pady=10)

        button5 = Button(window, text="RPS 11", font=customFont, bg="lightgray")
        button5.pack(fill=X, padx=180, pady=10)

        button6 = Button(window, text="RPS 15", font=customFont, bg="lightgray")
        button6.pack(fill=X, padx=180, pady=10)

        button7 = Button(window, text="RPS 25", font=customFont, bg="lightgray")
        button7.pack(fill=X, padx=180, pady=10)

        label = Label(window, text="RPS 101 coming never", font=customFont, fg="lightgray")
        label.pack(fill=X, pady=10)

    def rpsWindow(self):
        # create rps window
        rpsWindow = Toplevel()

        # modify window
        rpsWindow.iconbitmap("resources/Rock.ico")
        rpsWindow.title("classic RPS")
        rpsWindow.geometry(str(width) + "x" + str(height))
        rpsWindow.resizable(0, 0)

        rps = RPS(rpsWindow)

        rpsWindow.mainloop()

    def rpsslWindow(self):
        # create rpssl window
        rpsslWindow = Toplevel()

        # modify window
        rpsslWindow.iconbitmap("resources/Rock.ico")
        rpsslWindow.title("Rock, paper, scissors, spock, lizard")
        rpsslWindow.geometry(str(width) + "x" + str(height))
        rpsslWindow.resizable(0, 0)

        rps = RPSSL(rpsslWindow)

        rpsslWindow.mainloop()

class RPS(Frame):
    options = ("rock", "paper", "scissors")

    #computer:  # Rock                                  # paper                                     # scissors
    outcomes = (("You both played rock! It's a tie!",   "Paper covers rock! The computer wins!",     "Rock crushes scissors! You win!"), #R
                ("Paper covers rock! You win!",     "You both played paper! It's a tie!",       "Scissors cut paper! The computer wins!"), #P
                ("Rock crushes scissors! The computer wins!", "Scissors cut paper! You win!",   "You both played scissors! It's a tie!")) #S

    def __init__(self, parent):
        Frame.__init__(self, parent)

        self.parent = parent
        self.initUI()

    def initUI(self):
        # create font
        customFont = tkFont.Font(window, family="Helvetica", size=14)

        # image
        rpsImg = Image.open("resources/rps.png")
        rpsImage = ImageTk.PhotoImage(rpsImg, master=self.parent)

        # widgets
        image = Label(self.parent, image = rpsImage)
        image.image = rpsImage
        image.pack(side = LEFT)

        button1 = Button(self.parent, text = "Rock", font = customFont)
        button1.pack(fill = X, pady = 10, padx = 16)

        button2 = Button(self.parent, text = "Paper", font = customFont)
        button2.pack(fill = X, pady = 10, padx = 16)

        button3 = Button(self.parent, text = "Scissors", font = customFont)
        button3.pack(fill = X, pady = 10, padx = 16)

        scroll = Scrollbar(self.parent)
        scroll.pack(side=RIGHT, fill=Y, pady = 10)

        log = Text(self.parent, state = DISABLED)
        log.pack(fill = BOTH, pady = 10)

        scroll.config(command = log.yview)
        log.config(yscrollcommand = scroll.set)

        log.config(state=NORMAL)
        log.insert(END, "Pick an option!\n\n")
        log.config(state=DISABLED)

        button1.config(command = lambda: self.game(log, 0))
        button2.config(command = lambda: self.game(log, 1))
        button3.config(command = lambda: self.game(log, 2))

    def game(self, log, playerChoice):

        log.config(state=NORMAL)
        log.insert(END, "You played " + self.options[playerChoice] + "!\n")
        log.config(state=DISABLED)

        computerChoice = random.randint(0,2)

        log.config(state=NORMAL)
        log.insert(END, "The computer played " + self.options[computerChoice] + "!\n")
        log.config(state=DISABLED)

        log.config(state=NORMAL)
        log.insert(END, self.outcomes[playerChoice][computerChoice] + "\n\n")
        log.config(state=DISABLED)

        log.config(state=NORMAL)
        log.insert(END, "Pick an option!\n\n")
        log.config(state=DISABLED)


class RPSSL(Frame):
    options = ("rock", "paper", "scissors", "Spock", "lizard")

    ws = getWorksheet("resources/rpssl.xlsx")

    def __init__(self, parent):
        Frame.__init__(self, parent)

        self.parent = parent
        self.initUI()

    def initUI(self):
        # create font
        customFont = tkFont.Font(window, family="Helvetica", size=12)

        # image
        rpsImg = Image.open("resources/rock paper scissors spock lizard.png")
        rpsImage = ImageTk.PhotoImage(rpsImg, master=self.parent)

        # widgets
        image = Label(self.parent, image=rpsImage)
        image.image = rpsImage
        image.pack(side=LEFT)

        button1 = Button(self.parent, text="Rock", font=customFont)
        button1.pack(fill=X, pady=5, padx=16)

        button2 = Button(self.parent, text="Paper", font=customFont)
        button2.pack(fill=X, pady=5, padx=16)

        button3 = Button(self.parent, text="Scissors", font=customFont)
        button3.pack(fill=X, pady=5, padx=16)

        button4 = Button(self.parent, text="Spock", font=customFont)
        button4.pack(fill=X, pady=5, padx=16)

        button5 = Button(self.parent, text="Lizard", font=customFont)
        button5.pack(fill=X, pady=5, padx=16)

        scroll = Scrollbar(self.parent)
        scroll.pack(side=RIGHT, fill=Y, pady=5)

        log = Text(self.parent, state=DISABLED)
        log.pack(fill=BOTH, pady=5)

        scroll.config(command=log.yview)
        log.config(yscrollcommand=scroll.set)

        log.config(state=NORMAL)
        log.insert(END, "Pick an option!\n\n")
        log.config(state=DISABLED)

        button1.config(command=lambda: self.game(log, 0))
        button2.config(command=lambda: self.game(log, 1))
        button3.config(command=lambda: self.game(log, 2))
        button4.config(command=lambda: self.game(log, 3))
        button5.config(command=lambda: self.game(log, 4))

    def game(self, log, playerChoice):

        log.config(state=NORMAL)
        log.insert(END, "You played " + self.options[playerChoice] + "!\n")
        log.config(state=DISABLED)

        computerChoice = random.randint(0, 4)

        log.config(state=NORMAL)
        log.insert(END, "The computer played " + self.options[computerChoice] + "!\n")
        log.config(state=DISABLED)

        log.config(state=NORMAL)
        log.insert(END, self.ws.cell(row = playerChoice + 1, column = computerChoice + 1).value + "\n\n") # self.outcomes[playerChoice][computerChoice]
        log.config(state=DISABLED)

        log.config(state=NORMAL)
        log.insert(END, "Pick an option!\n\n")
        log.config(state=DISABLED)


### main window ###

# create window
window = Tk()

# modify window
window.iconbitmap("resources/Rock.ico")
window.title("Rock, paper, scissors: Extended")
window.geometry(str(width) + "x" + str(height))
window.resizable(0, 0)

app = Application(window)

# start main loop
window.mainloop()
